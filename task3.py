from openpyxl import load_workbook

from ortools.sat.python import cp_model

model = cp_model.CpModel()

# Load the XLSX file
xlsx_file = load_workbook('Assignment_DA_1_data.xlsx', read_only=True)

projects_sheet = xlsx_file['Projects']
quotes_sheet = xlsx_file['Quotes']
dependencies_sheet = xlsx_file['Dependencies']
value_sheet = xlsx_file['Value']

data = []
for row in projects_sheet.iter_rows(values_only=True):
    data.append([str(cell) for cell in row])

months = data[0][1:]
projects = [x[0] for x in data[1:]]
jobs_projects = {x[0] : [(job, m) for job, m in zip(x[1:], months) if job != 'None'] for x in  data[1: ]}

data = []
for row in quotes_sheet.iter_rows(values_only=True):
    data.append([str(cell) for cell in row])

jobs = data[0][1:]
contractors = [x[0] for x in data[1:]]
quotes = {}
job_contractors = {}
for j in range(len(jobs)):
    job_contractors[jobs[j]] = []
    for i in range(len(contractors)):
        if data[i + 1][j + 1] != 'None':
            quotes[(contractors[i], jobs[j])] = int(data[i + 1][j + 1])
            job_contractors[jobs[j]].append(contractors[i])
        else:
            quotes[(contractors[i], jobs[j])] = 0

data = []
for row in dependencies_sheet.iter_rows(values_only=True):
    data.append([str(cell) for cell in row])

dependencies = {}
for i in range(len(projects)):
    for j in range(len(projects)):
            dependencies[(projects[i], projects[j])] = data[i + 1][j + 1]

data = []
for row in value_sheet.iter_rows(values_only=True):
    data.append([str(cell) for cell in row])

values = {p: int(x) for (p, x) in data[1:]}

# projects vars
projects_vars = {}
for p in projects:
    for j in jobs:
        for m in months:
            for c in contractors:
                projects_vars[p, j, c, m] = model.NewBoolVar('projects_vars[%s, %s, %s, %s]' % (p, j, c, m))

projects_selected_vars = {p: model.NewBoolVar('') for p in projects}

# only quoted contractors can do a certain job at a certain month
for p in projects:
    for j in jobs:
        for m in months:
            for c in contractors:
                if c not in job_contractors[j]:
                    model.Add(projects_vars[p, j, c, m] == 0)
                if (j, m) not in jobs_projects[p]:
                    model.Add(projects_vars[p, j, c, m] == 0)

# if project is selected all the jobs for the projects must be done by exactly one contractor
for p in projects:
    for (j, m) in jobs_projects[p]:
        model.Add(projects_selected_vars[p] == sum(projects_vars[p, j, c, m] for c in contractors))

# one contractor can not do multiple jobs in the same month
for c in contractors:
    for m in months:
        model.Add(sum(projects_vars[p, j, c, m] for j in jobs for p in projects) <= 1)


# conflicts
for p1 in projects:
    for p2 in projects:
        if dependencies[(p1, p2)] == 'required':
            model.AddImplication(projects_selected_vars[p1], projects_selected_vars[p2])
        if dependencies[(p1, p2)] == 'conflict':
            model.AddImplication(projects_selected_vars[p1], projects_selected_vars[p2].Not())

# profit must exceed 2160
profit = sum(values[p] * projects_selected_vars[p] for p in projects) - sum(projects_vars[p, j, c, m] * quotes[c, j] for p in projects for j in jobs for c in contractors for m in months)
model.Add(profit >= 2160)

class SolutionPrinter(cp_model.CpSolverSolutionCallback):
    def __init__(self, variables, objective):
        cp_model.CpSolverSolutionCallback.__init__(self)
        self.__variables = variables
        self.__objective = objective
        self.__solution_count = 0

    def on_solution_callback(self):
        self.__solution_count += 1
        with open('result_3.txt', 'a') as f:
            f.write('Solution %i, profit = %i\n' %
                  (self.__solution_count, self.Value(self.__objective)))
            for p in projects:
                if self.Value(projects_selected_vars[p]):
                   f.write(f'  {p}:\n')
                   for j, m in jobs_projects[p]:
                       for c in contractors:
                            if self.Value(projects_vars[p, j, c, m]):
                                f.write(f'    {j}: {c} at {m}\n')
            f.write('\n')
    def solution_count(self):
        return self.__solution_count

with open('result_3.txt', 'w') as f:
    f.write('')

solver = cp_model.CpSolver()
solution_printer = SolutionPrinter(list(projects_vars.values()) + list(projects_selected_vars.values()), profit)
status = solver.SearchForAllSolutions(model, solution_printer)

print(f"Total Solutions: {solution_printer.solution_count()}")
with open('result_3.txt', 'a') as f:
    f.write(f"Total Solutions: {solution_printer.solution_count()}\n")
